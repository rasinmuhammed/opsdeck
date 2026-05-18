"""
CSV/Excel Export functionality for FastAPI Shadcn Admin.

Provides utilities to export model data to CSV and Excel formats.
"""

from __future__ import annotations

import csv
import io
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from sqlalchemy.ext.asyncio import AsyncSession

__all__ = ["CSVExporter", "ExcelExporter", "export_to_csv", "export_to_excel"]


class CSVExporter:
    """
    CSV export utility for model data.

    Features:
    - Exports query results to CSV format
    - Handles pagination
    - Configurable field selection
    - Proper encoding (UTF-8 BOM for Excel compatibility)

    Usage:
        exporter = CSVExporter()
        csv_content = await exporter.export(session, User, fields=["id", "email"])
    """

    def __init__(self, encoding: str = "utf-8-sig"):
        """
        Initialize CSV exporter.

        Args:
            encoding: Character encoding (utf-8-sig adds BOM for Excel)
        """
        self.encoding = encoding

    async def export(
        self,
        session: "AsyncSession",
        model: type,
        fields: list[str] | None = None,
        filters: dict[str, Any] | None = None,
        max_rows: int = 10000,
    ) -> bytes:
        """
        Export model data to CSV.

        Args:
            session: Database session
            model: SQLAlchemy model to export
            fields: Fields to include (None = all)
            filters: Optional filters to apply
            max_rows: Maximum rows to export (safety limit)

        Returns:
            CSV content as bytes
        """
        from sqlalchemy import select

        # Build query
        query = select(model)
        if filters:
            for key, value in filters.items():
                query = query.where(getattr(model, key) == value)

        # Limit for safety
        query = query.limit(max_rows)

        # Execute
        result = await session.execute(query)
        rows = result.scalars().all()

        # Determine fields
        if fields is None:
            # Get all column names
            fields = [c.name for c in model.__table__.columns]

        # Generate CSV
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=fields)
        writer.writeheader()

        for row in rows:
            row_dict = {field: getattr(row, field, None) for field in fields}
            writer.writerow(row_dict)

        # Return as bytes
        return output.getvalue().encode(self.encoding)


async def export_to_csv(
    session: "AsyncSession",
    model: type,
    fields: list[str] | None = None,
    **kwargs,
) -> bytes:
    """Convenience function to export model to CSV."""
    exporter = CSVExporter()
    return await exporter.export(session, model, fields=fields, **kwargs)


class ExcelExporter:
    """
    Excel (.xlsx) export utility for model data.

    Requires openpyxl: pip install openpyxl
    Falls back gracefully with an ImportError if not installed.
    """

    async def export(
        self,
        session: "AsyncSession",
        model: type,
        fields: list[str] | None = None,
        filters: dict[str, Any] | None = None,
        max_rows: int = 10000,
        sheet_name: str = "Data",
    ) -> bytes:
        """
        Export model data to Excel (.xlsx).

        Args:
            session: Database session
            model: SQLAlchemy model to export
            fields: Fields to include (None = all columns)
            filters: Optional key=value filters to apply
            max_rows: Safety row limit
            sheet_name: Name of the Excel sheet

        Returns:
            Excel file content as bytes

        Raises:
            ImportError: If openpyxl is not installed
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            ) from exc

        from sqlalchemy import select

        query = select(model)
        if filters:
            for key, value in filters.items():
                query = query.where(getattr(model, key) == value)
        query = query.limit(max_rows)

        result = await session.execute(query)
        rows = result.scalars().all()

        if fields is None:
            fields = [c.name for c in model.__table__.columns]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Header row with styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")

        for col_idx, field in enumerate(fields, start=1):
            cell = ws.cell(row=1, column=col_idx, value=field.replace("_", " ").title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            ws.column_dimensions[cell.column_letter].width = max(len(field) + 4, 12)

        # Data rows
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, field in enumerate(fields, start=1):
                value = getattr(row, field, None)
                if hasattr(value, "isoformat"):
                    value = value.isoformat()
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Freeze the header row
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


async def export_to_excel(
    session: "AsyncSession",
    model: type,
    fields: list[str] | None = None,
    **kwargs,
) -> bytes:
    """Convenience function to export model to Excel."""
    exporter = ExcelExporter()
    return await exporter.export(session, model, fields=fields, **kwargs)

"""
Admin Router for FastAPI Shadcn Admin.

Implements all admin endpoints with security validation:
- Dashboard and list views
- Create, edit, delete operations
- Signed fragment loading for polymorphic forms
"""

from __future__ import annotations

import math
import platform
import psutil
import os
from datetime import datetime, timedelta
from typing import Any, TYPE_CHECKING, Callable, AsyncGenerator

from fastapi import APIRouter, Request, HTTPException, Depends, Query
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse
from starlette.status import (
    HTTP_401_UNAUTHORIZED,
    HTTP_403_FORBIDDEN,
    HTTP_404_NOT_FOUND,
)

if TYPE_CHECKING:
    from fastapi_matrix_admin.core.registry import AdminRegistry
    from fastapi_matrix_admin.core.security import URLSigner
    from fastapi_matrix_admin.core.integrator import SchemaWalker
    from sqlalchemy.ext.asyncio import AsyncSession

from fastapi_matrix_admin.core.crud import CRUDBase
from fastapi_matrix_admin.core.integrator import FieldDefinition, FieldType
from fastapi_matrix_admin.audit.models import AuditLogger
from fastapi_matrix_admin.auth.models import PermissionChecker
from fastapi_matrix_admin.auth.service import AuthService
from fastapi_matrix_admin.core.rate_limiter import RateLimiter


def extract_sqlalchemy_fields(
    model: Any,
    exclude: list[str] | None = None,
    include: list[str] | None = None,
    registry: "AdminRegistry | None" = None,
    field_overrides: dict[str, dict[str, Any]] | None = None,
    widgets: dict[str, str] | None = None,
) -> list[FieldDefinition]:
    """
    Extract fields from a SQLAlchemy model and convert to FieldDefinition objects.

    Args:
        model: SQLAlchemy model class
        exclude: List of field names to exclude
        include: List of field names to include (if specified, only these fields are included)

    Returns:
        List of FieldDefinition objects for template rendering
    """
    from sqlalchemy import inspect as sqla_inspect
    from sqlalchemy.orm import ColumnProperty, RelationshipProperty

    mapper = sqla_inspect(model)
    fields = []

    # Get all columns
    for attr in mapper.attrs:
        if isinstance(attr, RelationshipProperty) and attr.secondary is not None:
            # Many-to-many relationship via secondary table
            field_name = attr.key
            if include and field_name not in include:
                continue
            if exclude and field_name in exclude:
                continue

            target_class = attr.mapper.class_
            target_name = target_class.__name__

            # Try to resolve to registry name
            if registry:
                for config in registry.all():
                    if config.model is target_class:
                        target_name = config.name
                        break

            field_def = FieldDefinition(
                name=field_name,
                field_type=FieldType.MANY_TO_MANY,
                required=False,
                title=field_name.replace("_", " ").title(),
                target_model=target_name,
            )
            fields.append(field_def)
            continue

        if isinstance(attr, ColumnProperty):
            column = attr.columns[0]
            field_name = attr.key

            # Apply include/exclude filters
            if include and field_name not in include:
                continue
            if exclude and field_name in exclude:
                continue

            # Skip primary key (id) - it's auto-generated
            if column.primary_key:
                continue

            # Determine field type from SQLAlchemy column type
            field_type = FieldType.TEXT  # Default
            python_type = column.type.python_type
            target_model = None

            # Check for Foreign Keys
            if column.foreign_keys:
                field_type = FieldType.RELATIONSHIP
                # Get target table name
                fk = list(column.foreign_keys)[0]
                # Try to map table name back to model name if possible, or just use table name
                # Optimally, we want the Model Name (e.g. "User") for the API search
                # We can store the table name and let the registry lookup handle it later?
                # Or simplistic approach: Use the table name.
                target_table = fk.column.table.name
                # TODO: We need a way to map table_name -> Registry Model Name
                # For now, let's assume table name or look up via registry reverse lookup in future.
                target_model = target_table
            elif python_type is bool:
                field_type = FieldType.BOOLEAN
            elif python_type is int:
                field_type = FieldType.NUMBER
            elif python_type is float:
                field_type = FieldType.FLOAT
            elif hasattr(column.type, "__visit_name__"):
                if column.type.__visit_name__ == "text":
                    field_type = FieldType.TEXTAREA
                elif column.type.__visit_name__ in ("json", "jsonb"):
                    field_type = FieldType.JSON
                elif column.type.__visit_name__ in ("date", "datetime"):
                    field_type = FieldType.DATETIME

            if registry and target_model:
                for config in registry.all():
                    if (
                        hasattr(config.model, "__table__")
                        and config.model.__table__.name == target_model
                    ):
                        target_model = config.name
                        break

            if widgets and widgets.get(field_name) == "json":
                field_type = FieldType.JSON

            # Create FieldDefinition
            field_def = FieldDefinition(
                name=field_name,
                field_type=field_type,
                required=not column.nullable,
                default=(
                    column.default.arg
                    if column.default and hasattr(column.default, "arg")
                    else None
                ),
                title=field_name.replace("_", " ").title(),
                description=None,
                placeholder=None,
                target_model=target_model,
            )
            overrides = (field_overrides or {}).get(field_name, {})
            for key, value in overrides.items():
                if hasattr(field_def, key):
                    setattr(field_def, key, value)

            fields.append(field_def)

    return fields


def model_to_dict(model: Any, exclude: set[str] | None = None) -> dict[str, Any]:
    """Serialize a SQLAlchemy or Pydantic model into a plain dict."""
    exclude = exclude or set()
    data: dict[str, Any] = {}
    if hasattr(model, "__table__"):
        for column in model.__table__.columns:
            if column.name in exclude:
                continue
            value = getattr(model, column.name, None)
            if isinstance(value, datetime):
                value = value.isoformat()
            data[column.name] = value
    elif hasattr(model, "model_dump"):
        data = model.model_dump(exclude=exclude)
    return data


async def _load_m2m_choices(
    session: Any,
    fields: list,
    registry: Any,
) -> dict[str, list[tuple[str, str]]]:
    """Load (id, label) choices for all MANY_TO_MANY fields from the database."""
    from sqlalchemy import select as sa_select

    choices: dict[str, list[tuple[str, str]]] = {}
    for f in fields:
        if f.field_type != FieldType.MANY_TO_MANY:
            continue
        target_model_class = None
        for config in registry.all():
            if config.name == f.target_model or config.model.__name__ == f.target_model:
                target_model_class = config.model
                break
        if target_model_class is None:
            choices[f.name] = []
            continue
        result = await session.execute(sa_select(target_model_class))
        instances = result.scalars().all()
        field_choices: list[tuple[str, str]] = []
        for inst in instances:
            inst_id = str(getattr(inst, "id", ""))
            if hasattr(inst, "__admin_repr__"):
                label = inst.__admin_repr__()
            else:
                label = str(inst)
            field_choices.append((inst_id, label))
        choices[f.name] = field_choices
    return choices


def create_admin_router(
    registry: "AdminRegistry",
    signer: "URLSigner",
    walker: "SchemaWalker",
    templates,
    prefix: str = "/admin",
    title: str = "Admin",
    session_dependency: (
        Callable[[], AsyncGenerator["AsyncSession", None]] | None
    ) = None,
    audit_logger: AuditLogger | None = None,
    auth_model: Any | None = None,
    demo_mode: bool = False,
    secure_cookies: bool | None = None,
    theme: str = "matrix",
) -> APIRouter:
    # Initialize rate limiter (e.g., 5 login attempts per minute)
    login_limiter = RateLimiter(rate=5, per=60)
    """
    Create the admin router with all endpoints.

    Args:
        registry: The AdminRegistry with registered models
        signer: URLSigner for token validation
        walker: SchemaWalker for form field generation
        templates: Jinja2Templates instance
        prefix: URL prefix for admin routes
        title: Admin panel title
        session_dependency: Optional async session dependency for database operations

    Returns:
        Configured APIRouter
    """
    router = APIRouter(tags=["admin"])
    auth_service = (
        AuthService(signer, auth_model)
        if auth_model is not None and session_dependency is not None
        else None
    )
    secure_cookie_flag = (
        secure_cookies
        if secure_cookies is not None
        else os.getenv("ADMIN_SECURE_COOKIES", "false").lower() == "true"
    )

    async def get_current_user(
        request: Request,
        session: "AsyncSession" | None,
    ) -> Any | None:
        if not auth_service or not session:
            return None
        return await auth_service.get_current_user(request, session)

    async def require_user(
        request: Request,
        session: "AsyncSession" | None,
    ) -> Any | None:
        if not auth_service:
            return None
        if session is None:
            raise HTTPException(
                status_code=HTTP_401_UNAUTHORIZED,
                detail="Authentication requires a database session.",
            )
        user = await auth_service.get_current_user(request, session)
        if user is None:
            raise HTTPException(
                status_code=HTTP_401_UNAUTHORIZED,
                detail="Not authenticated",
                headers={"WWW-Authenticate": "Cookie"},
            )
        return user

    async def enforce_permission(
        request: Request,
        session: "AsyncSession" | None,
        model_config: Any | None,
        permission: str,
    ) -> Any | None:
        if model_config is None:
            return await require_user(request, session) if auth_service else None

        if not auth_service:
            return None

        user = await require_user(request, session)
        checker = PermissionChecker(user, model_config.permissions)
        allowed = getattr(checker, f"can_{permission}", checker.can_view)()
        if not allowed:
            raise HTTPException(
                status_code=HTTP_403_FORBIDDEN,
                detail=f"Not allowed to {permission} {model_config.name}",
            )
        return user

    def scoped_query_transform(
        model_config: Any, request: Request, session: Any, user: Any
    ):
        if not model_config or not model_config.row_scope:
            return None

        def _apply(query: Any) -> Any:
            return model_config.row_scope(
                request=request, query=query, session=session, user=user
            )

        return _apply

    def visible_actions(model_config: Any, request: Request, user: Any) -> list[Any]:
        actions = []
        for action in model_config.actions:
            if action.visible is None or action.visible(request=request, user=user):
                actions.append(action)
        return actions

    def render_action_result(result: Any, fallback_url: str) -> Any:
        if isinstance(result, (HTMLResponse, JSONResponse, RedirectResponse)):
            return result
        if isinstance(result, dict):
            return JSONResponse(result)
        return RedirectResponse(url=fallback_url, status_code=303)

    _base_template = (
        "layouts/base_clean.html" if theme == "clean" else "layouts/base.html"
    )

    def get_common_context(request: Request) -> dict[str, Any]:
        """Get common template context."""
        return {
            "request": request,
            "admin_title": title,
            "base_template": _base_template,
            "models": sorted(
                registry.all(),
                key=lambda config: (
                    config.menu_order,
                    config.menu_label or config.name,
                ),
            ),
            "csp_nonce": getattr(request.state, "csp_nonce", ""),
        }

    # ==================== Auth ====================

    @router.get("/login", response_class=HTMLResponse, name="admin:login")
    async def login_view(request: Request):
        """Show login form."""
        context = {
            "request": request,
            "title": title,
            "prefix": prefix,
            "csp_nonce": getattr(request.state, "csp_nonce", ""),
            "error": request.query_params.get("error"),
            "demo_mode": demo_mode,
        }
        return templates.TemplateResponse(request, "pages/login.html", context)

    @router.post("/login", response_class=HTMLResponse, name="admin:login_submit")
    async def login_submit(
        request: Request,
        session: "AsyncSession" = (
            Depends(session_dependency) if session_dependency else None
        ),
    ):
        """Handle login submission with rate limiting and security."""
        # 1. IP-based Rate Limiting
        client_ip = request.client.host if request.client else "unknown"
        if not login_limiter.consume(client_ip):
            # Slow down response to prevent timing attacks
            import asyncio

            await asyncio.sleep(1.0)

            context = {
                "request": request,
                "title": title,
                "prefix": prefix,
                "csp_nonce": getattr(request.state, "csp_nonce", ""),
                "error": "Too many login attempts. Please try again later.",
            }
            return templates.TemplateResponse(
                request, "pages/login.html", context, status_code=429
            )

        form_data = await request.form()
        username = form_data.get("username")
        password = form_data.get("password")

        user = None
        valid_password = False

        # 2. Database Lookup & Auth Logic
        if session and auth_model:
            # Bulletproof DB Auth
            from sqlalchemy import select

            stmt = select(auth_model).where(auth_model.username == username)
            result = await session.execute(stmt)
            user = result.scalar_one_or_none()

            if user and hasattr(user, "verify_password"):
                valid_password = user.verify_password(password)

        # If no user found via DB lookups then fails.
        if not user or not valid_password:
            # Add a small delay for failed attempts
            import asyncio

            await asyncio.sleep(0.5)

            return templates.TemplateResponse(
                request,
                "pages/login.html",
                {
                    "request": request,
                    "title": title,
                    "prefix": prefix,
                    "csp_nonce": getattr(request.state, "csp_nonce", ""),
                    "error": "Invalid credentials",
                },
                status_code=401,
            )

        if hasattr(user, "is_active") and not user.is_active:
            return templates.TemplateResponse(
                request,
                "pages/login.html",
                {
                    "request": request,
                    "admin_title": title,
                    "csp_nonce": getattr(request.state, "csp_nonce", ""),
                    "error": "Account disabled",
                },
                status_code=403,
            )

        # 4. Session Creation
        from fastapi_matrix_admin.auth.models import SessionData

        session_data = SessionData.create(
            user, remember_me=bool(form_data.get("remember_me"))
        )

        # Serialize session (using signer for tamper-proof cookie)
        session_token = signer.sign(session_data.model_dump())

        # 5. Success - Set Cookie & Redirect
        response = RedirectResponse(url=f"{prefix}/", status_code=303)
        response.set_cookie(
            key="admin_session",
            value=session_token,
            httponly=True,
            samesite="lax",
            secure=secure_cookie_flag,
            max_age=3600 * 24 * 30 if session_data.expires_at else 3600 * 24,
        )

        # Audit Log Login
        if audit_logger:
            # We don't await this to keep login fast? No, await for reliability.
            # Using fire-and-forget might be better for perf but safety first.
            await audit_logger.log_create(
                session,
                "Auth",
                "login",
                {"username": username},
                user_id=user.id,
                ip_address=client_ip,
            )

        return response

    @router.get("/logout", name="admin:logout")
    async def logout(request: Request):
        response = RedirectResponse(url=f"{prefix}/login", status_code=303)
        response.delete_cookie("admin_session")
        return response

    # ==================== Dashboard ====================

    @router.get("/", response_class=HTMLResponse, name="admin:index")
    async def dashboard(
        request: Request,
        session: "AsyncSession" = (
            Depends(session_dependency) if session_dependency else None
        ),
    ):
        """Admin dashboard with analytics."""
        from datetime import datetime
        from sqlalchemy import func, select

        current_user = await enforce_permission(request, session, None, "view")

        # Get registered models
        registered_models = registry.get_all()  # Returns list of model names

        # Calculate KPIs
        kpis = []
        model_counts = {}

        if session and session_dependency:
            # Count records per model
            for model_name in registered_models:
                try:
                    config = registry.get(model_name)
                    if hasattr(config.model, "__tablename__"):
                        result = await session.execute(
                            select(func.count()).select_from(config.model)
                        )
                        count = result.scalar() or 0
                        model_counts[model_name] = count
                except Exception:
                    model_counts[model_name] = 0

            # Total records KPI
            total_records = sum(model_counts.values())
            kpis.append(
                {
                    "label": "Total Records",
                    "value": f"{total_records:,}",
                    "icon": "database",
                    "change": None,
                }
            )

            # Models KPI
            kpis.append(
                {
                    "label": "Models",
                    "value": len(registered_models),
                    "icon": "layers",
                    "change": None,
                }
            )
        else:
            # No database - show model count only
            kpis.append(
                {
                    "label": "Models",
                    "value": len(registered_models),
                    "icon": "layers",
                    "change": None,
                }
            )

        # Chart data (sample data if no database)
        # Chart data (sample data if no database)
        activity_labels = []
        activity_values = []

        from sqlalchemy import desc

        if session and audit_logger:
            # Query actual activity from audit log
            for i in range(6, -1, -1):
                date = datetime.now() - timedelta(days=i)
                activity_labels.append(date.strftime("%a"))

                start_date = date.replace(hour=0, minute=0, second=0, microsecond=0)
                end_date = start_date + timedelta(days=1)

                stmt = (
                    select(func.count())
                    .select_from(audit_logger.audit_model)
                    .where(
                        audit_logger.audit_model.created_at >= start_date,
                        audit_logger.audit_model.created_at < end_date,
                    )
                )
                res = await session.execute(stmt)
                count = res.scalar() or 0
                activity_values.append(count)
        else:
            # Last 7 days sample
            for i in range(6, -1, -1):
                date = datetime.now() - timedelta(days=i)
                activity_labels.append(date.strftime("%a"))
                activity_values.append(max(0, 10 - i * 2 + (i % 3)))

        # Model distribution data
        model_labels = list(model_counts.keys())[:5]  # Top 5 models
        model_values = [model_counts.get(name, 0) for name in model_labels]

        # If no data, use sample
        if not model_values or sum(model_values) == 0:
            model_labels = registered_models[:5]
            model_values = [5, 3, 2, 1, 1]

        chart_data = {
            "activity_labels": activity_labels,
            "activity_values": activity_values,
            "model_labels": model_labels,
            "model_values": model_values,
        }

        # Recent activity (if audit logging enabled)
        recent_activity = []
        if session and audit_logger:
            from sqlalchemy import desc

            stmt = (
                select(audit_logger.audit_model)
                .order_by(desc(audit_logger.audit_model.created_at))
                .limit(10)
            )
            result = await session.execute(stmt)
            logs = result.scalars().all()
            for log in logs:
                action_str = (
                    log.action.value
                    if hasattr(log.action, "value")
                    else str(log.action)
                )
                recent_activity.append(
                    {
                        "action": action_str,
                        "model_name": log.model_name,
                        "username": log.username or "SYSTEM",
                        "created_at": log.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                    }
                )

        # System KPIs (Observer Module)
        try:
            boot_time = datetime.fromtimestamp(psutil.boot_time()).strftime(
                "%Y-%m-%d %H:%M:%S"
            )
        except (PermissionError, OSError):
            boot_time = "Unavailable"

        try:
            _disk_path = "C:\\" if platform.system() == "Windows" else "/"
            disk_usage_pct = psutil.disk_usage(_disk_path).percent
        except (PermissionError, OSError):
            disk_usage_pct = 0

        system_stats = {
            "platform": f"{platform.system()} {platform.release()}",
            "processor": platform.processor(),
            "cpu_usage": psutil.cpu_percent(interval=None),
            "ram_usage": psutil.virtual_memory().percent,
            "ram_total": f"{round(psutil.virtual_memory().total / (1024**3), 2)} GB",
            "disk_usage": disk_usage_pct,
            "boot_time": boot_time,
        }

        # Add System Stats to KPIs
        kpis.append(
            {
                "label": "CPU Load",
                "value": f"{system_stats['cpu_usage']}%",
                "icon": "cpu",
                "change": None,
            }
        )
        kpis.append(
            {
                "label": "RAM Usage",
                "value": f"{system_stats['ram_usage']}%",
                "icon": "activity",
                "change": None,
            }
        )

        context = {
            **get_common_context(request),
            "kpis": kpis,
            "registered_models": registered_models,
            "chart_data": chart_data,
            "recent_activity": recent_activity,
            "system_stats": system_stats,  # Pass full stats if needed
            "current_user": current_user,
        }

        return templates.TemplateResponse(request, "pages/index.html", context)

    # ==================== List View ====================

    @router.get("/{model}/", response_class=HTMLResponse, name="admin:list")
    async def list_view(
        request: Request,
        model: str,
        page: int = Query(1, ge=1),
        per_page: int = Query(25, ge=1, le=100),
        search: str | None = Query(None),
        session: "AsyncSession" = (
            Depends(session_dependency) if session_dependency else None
        ),
    ):
        """List all records of a model."""
        # Validate model access
        try:
            model_config = registry.validate_model_access(model)
        except Exception:
            raise HTTPException(
                status_code=HTTP_403_FORBIDDEN, detail="Model not registered"
            )

        current_user = await enforce_permission(request, session, model_config, "view")
        query_transform = scoped_query_transform(
            model_config, request, session, current_user
        )

        # Get field definitions for column headers
        # For SQLAlchemy models, skip schema walking (use list_display directly)
        if hasattr(model_config.model, "__tablename__"):
            # SQLAlchemy model - use list_display directly
            display_fields = model_config.list_display or ["id"]
        else:
            # Pydantic model - use schema walker
            fields = walker.walk(model_config.model)
            display_fields = model_config.list_display or [f.name for f in fields[:5]]

        columns = [
            {"field": name, "label": name.replace("_", " ").title(), "sortable": True}
            for name in display_fields
        ]

        # Prepare Filter Definitions
        # We need to know the type of each filter field to render the correct widget
        filter_definitions = []
        if model_config.filter_fields:
            # Create a lookup map only for fields in filter_fields
            # If using SQLAlchemy, we might need to be clever.
            # Reuse the extraction logic or walker logic.

            all_fields = []
            if hasattr(model_config.model, "__tablename__"):
                all_fields = extract_sqlalchemy_fields(
                    model_config.model,
                    include=model_config.filter_fields,
                    registry=registry,
                    field_overrides=model_config.field_overrides,
                    widgets=model_config.widgets,
                )
            else:
                all_fields = walker.walk(
                    model_config.model, include=model_config.filter_fields
                )

            # Re-order to match filter_fields configuration
            field_map = {f.name: f for f in all_fields}
            for name in model_config.filter_fields:
                if name in field_map:
                    filter_definitions.append(field_map[name])

        # Extract filters from query params
        reserved_params = {"page", "per_page", "search"}
        filters = {
            k: v for k, v in request.query_params.items() if k not in reserved_params
        }

        # Fetch data from database if session available
        if session and hasattr(model_config.model, "__tablename__"):
            # SQLAlchemy model - use CRUD
            crud = CRUDBase(model_config.model)

            # Optimization: Detect relationships to eager load (N+1 prevention)
            load_relationships = []
            from sqlalchemy import inspect

            inspector = inspect(model_config.model)
            for field in display_fields:
                if field in inspector.relationships:
                    load_relationships.append(field)

            rows_data, total = await crud.list(
                session,
                page=page,
                per_page=per_page,
                search=search,
                search_fields=model_config.searchable_fields,
                filters=filters,
                order_by=model_config.ordering,
                load_relationships=sorted(
                    set(load_relationships + (model_config.eager_load or []))
                ),
                query_transform=query_transform,
            )
        else:
            # Pydantic/No-DB model - empty or mock data
            rows_data = []  # TODO: Maybe implement in-memory filter?
            total = 0

        # Context Variables
        edit_url_template = router.url_path_for("admin:edit", model=model, id="__id__")
        delete_url_template = router.url_path_for(
            "admin:delete", model=model, id="__id__"
        )
        delete_tokens = {
            str(getattr(row, "id", "")): signer.sign(
                {"model": model, "action": "delete"}
            )
            for row in rows_data
            if getattr(row, "id", None) is not None
        }

        context = {
            **get_common_context(request),
            "model": model,
            "model_config": model_config,
            "columns": columns,
            "rows": rows_data,
            "page": page,
            "per_page": per_page,
            "total": total,
            "total_pages": math.ceil(total / per_page) if per_page > 0 else 1,
            "search_query": search,
            "id_field": "id",
            "edit_url_template": edit_url_template,
            "delete_url_template": delete_url_template,
            "active_filters": filters,
            "filter_definitions": filter_definitions,
            "current_user": current_user,
            "actions": visible_actions(model_config, request, current_user),
            "delete_tokens": delete_tokens,
            "bulk_actions": ["bulk_delete"]
            + [
                action.name
                for action in visible_actions(model_config, request, current_user)
                if action.bulk
            ],
        }

        return templates.TemplateResponse(request, "pages/list.html", context)

    @router.get("/{model}/export/csv", name="admin:export_csv")
    async def export_csv(
        request: Request,
        model: str,
        search: str | None = Query(None),
        session: "AsyncSession" = (
            Depends(session_dependency) if session_dependency else None
        ),
    ):
        """Export filtered data as CSV."""
        try:
            model_config = registry.validate_model_access(model)
        except Exception:
            raise HTTPException(status_code=403, detail="Model not registered")

        current_user = await enforce_permission(
            request, session, model_config, "export"
        )
        query_transform = scoped_query_transform(
            model_config, request, session, current_user
        )

        # Reuse filter logic
        reserved_params = {"page", "per_page", "search"}
        filters = {
            k: v for k, v in request.query_params.items() if k not in reserved_params
        }

        # Determine fields to export
        if hasattr(model_config.model, "__tablename__"):
            fields = model_config.list_display or ["id"]  # Fallback
            # If list_display is empty, maybe get all columns?
            if not model_config.list_display:
                from sqlalchemy import inspection

                fields = [c.key for c in inspection.inspect(model_config.model).columns]
        else:
            fields = model_config.list_display
            if not fields:
                fields = [f.name for f in walker.walk(model_config.model)]

        async def generate_csv():
            import csv
            import io

            # Helper for writing valid CSV lines to stream
            def line(data):
                si = io.StringIO()
                cw = csv.writer(si)
                cw.writerow(data)
                return si.getvalue()

            # Yield Header
            yield line([f.replace("_", " ").title() for f in fields])

            if session and hasattr(model_config.model, "__tablename__"):
                crud = CRUDBase(model_config.model)
                page = 1
                while True:
                    # Fetch in chunks of 500
                    rows, _ = await crud.list(
                        session,
                        page=page,
                        per_page=500,
                        search=search,
                        search_fields=model_config.searchable_fields,
                        filters=filters,
                        order_by=model_config.ordering,
                        query_transform=query_transform,
                    )

                    if not rows:
                        break

                    for row in rows:
                        yield line([getattr(row, f, "") for f in fields])

                    page += 1
            else:
                # No DB support yet for non-SQLAlchemy export in this MVP
                yield line(["No database session available"])

        from fastapi.responses import StreamingResponse

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        filename = f"{model_config.name}_export_{timestamp}.csv"

        return StreamingResponse(
            generate_csv(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    @router.get("/{model}/export/xlsx", name="admin:export_xlsx")
    async def export_xlsx(
        request: Request,
        model: str,
        search: str | None = Query(None),
        session: "AsyncSession" = (
            Depends(session_dependency) if session_dependency else None
        ),
    ):
        """Export filtered data as Excel (.xlsx)."""
        try:
            model_config = registry.validate_model_access(model)
        except Exception:
            raise HTTPException(status_code=403, detail="Model not registered")

        current_user = await enforce_permission(
            request, session, model_config, "export"
        )
        query_transform = scoped_query_transform(
            model_config, request, session, current_user
        )

        reserved_params = {"page", "per_page", "search"}
        filters = {
            k: v for k, v in request.query_params.items() if k not in reserved_params
        }

        if hasattr(model_config.model, "__tablename__"):
            fields = model_config.list_display or [
                c.key for c in model_config.model.__table__.columns
            ]
        else:
            fields = model_config.list_display or [
                f.name for f in walker.walk(model_config.model)
            ]

        try:
            import openpyxl  # noqa: F401
        except ImportError:
            raise HTTPException(
                status_code=501,
                detail=(
                    "Excel export requires openpyxl. "
                    "Install it with: pip install openpyxl"
                ),
            )

        from fastapi.responses import Response

        all_rows: list[Any] = []

        if session and hasattr(model_config.model, "__tablename__"):
            crud = CRUDBase(model_config.model)
            page = 1
            while True:
                rows, _ = await crud.list(
                    session,
                    page=page,
                    per_page=500,
                    search=search,
                    search_fields=model_config.searchable_fields,
                    filters=filters,
                    order_by=model_config.ordering,
                    query_transform=query_transform,
                )
                if not rows:
                    break
                all_rows.extend(rows)
                page += 1

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            import io as _io

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = model_config.name[:31]

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="1E3A5F", end_color="1E3A5F", fill_type="solid"
            )
            for col_idx, field in enumerate(fields, start=1):
                cell = ws.cell(
                    row=1,
                    column=col_idx,
                    value=field.replace("_", " ").title(),
                )
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.column_dimensions[cell.column_letter].width = max(len(field) + 4, 12)

            for row_idx, row in enumerate(all_rows, start=2):
                for col_idx, field in enumerate(fields, start=1):
                    value = getattr(row, field, None)
                    if hasattr(value, "isoformat"):
                        value = value.isoformat()
                    ws.cell(row=row_idx, column=col_idx, value=value)

            ws.freeze_panes = "A2"
            buf = _io.BytesIO()
            wb.save(buf)
            xlsx_bytes = buf.getvalue()
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"Excel export failed: {exc}")

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        filename = f"{model_config.name}_export_{timestamp}.xlsx"

        return Response(
            content=xlsx_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    # ==================== Create View ====================

    @router.get("/{model}/create", response_class=HTMLResponse, name="admin:create")
    async def create_view(
        request: Request,
        model: str,
        session: "AsyncSession" = (
            Depends(session_dependency) if session_dependency else None
        ),
    ):
        """Show create form for a model."""
        try:
            model_config = registry.validate_model_access(model)
        except Exception:
            raise HTTPException(
                status_code=HTTP_403_FORBIDDEN, detail="Model not registered"
            )

        if model_config.readonly:
            raise HTTPException(
                status_code=HTTP_403_FORBIDDEN, detail="Model is read-only"
            )

        current_user = await enforce_permission(
            request, session, model_config, "create"
        )

        # Get form fields - use SQLAlchemy inspector for SQLAlchemy models
        if hasattr(model_config.model, "__tablename__"):
            # SQLAlchemy model
            fields = extract_sqlalchemy_fields(
                model_config.model,
                exclude=model_config.exclude,
                include=model_config.fields,
                registry=registry,
                field_overrides=model_config.field_overrides,
                widgets=model_config.widgets,
            )
        else:
            # Pydantic model
            fields = walker.walk(
                model_config.model,
                exclude=model_config.exclude,
                include=model_config.fields if model_config.fields else None,
            )

        # Generate signed fragment URL for polymorphic forms
        fragment_token = signer.create_fragment_token(model, action="load_fragment")
        fragment_url = f"{prefix}/fragments?token={fragment_token}"

        m2m_choices: dict = {}
        if session:
            m2m_choices = await _load_m2m_choices(session, fields, registry)

        context = {
            **get_common_context(request),
            "model_config": model_config,
            "current_model": model,
            "fields": fields,
            "values": {},
            "errors": {},
            "record_id": None,
            "fragment_url": fragment_url,
            "csrf_token": signer.sign({"action": "create", "model": model}),
            "current_user": current_user,
            "detail_panels": model_config.detail_panels,
            "delete_token": signer.sign({"model": model, "action": "delete"}),
            "m2m_choices": m2m_choices,
            "m2m_values": {},
        }

        return templates.TemplateResponse(request, "pages/edit.html", context)

    @router.post(
        "/{model}/create", response_class=HTMLResponse, name="admin:create_submit"
    )
    async def create_submit(
        request: Request,
        model: str,
        session: "AsyncSession" = (
            Depends(session_dependency) if session_dependency else None
        ),
    ):
        """Handle create form submission."""
        try:
            model_config = registry.validate_model_access(model)
        except Exception:
            raise HTTPException(
                status_code=HTTP_403_FORBIDDEN, detail="Model not registered"
            )

        if model_config.readonly:
            raise HTTPException(
                status_code=HTTP_403_FORBIDDEN, detail="Model is read-only"
            )

        current_user = await enforce_permission(
            request, session, model_config, "create"
        )
        scoped_query_transform(model_config, request, session, current_user)

        form_data = await request.form()

        # Validate CSRF token
        csrf_token = form_data.get("_csrf_token", "")
        try:
            signer.unsign(csrf_token)
        except Exception:
            raise HTTPException(
                status_code=HTTP_403_FORBIDDEN, detail="Invalid CSRF token"
            )

        # Create record in database if session available
        if session and hasattr(model_config.model, "__tablename__"):
            crud = CRUDBase(model_config.model)

            # Convert form data to dict (excluding internal fields)
            raw_data = {
                k: v for k, v in dict(form_data).items() if not k.startswith("_")
            }

            # Convert string values to appropriate types based on model
            create_data = {}
            m2m_fields_data: dict[str, list[str]] = {}
            fields = extract_sqlalchemy_fields(
                model_config.model,
                registry=registry,
                field_overrides=model_config.field_overrides,
                widgets=model_config.widgets,
            )
            for field in fields:
                if field.field_type == FieldType.MANY_TO_MANY:
                    m2m_fields_data[field.name] = form_data.getlist(field.name)
                    continue
                if field.name in raw_data:
                    value = raw_data[field.name]
                    # Convert boolean strings to actual booleans
                    if field.field_type == FieldType.BOOLEAN:
                        create_data[field.name] = value in (
                            "true",
                            "True",
                            "1",
                            "on",
                            True,
                        )
                    # Convert numeric strings to numbers
                    elif field.field_type == FieldType.NUMBER and isinstance(
                        value, str
                    ):
                        create_data[field.name] = int(value) if value else None
                    elif field.field_type == FieldType.FLOAT and isinstance(value, str):
                        create_data[field.name] = float(value) if value else None
                    elif field.field_type == FieldType.JSON and isinstance(value, str):
                        import json

                        create_data[field.name] = json.loads(value) if value else None
                    else:
                        create_data[field.name] = value if value != "" else None

            try:
                created_record = await crud.create(session, obj_in=create_data)

                # Set M2M relationships after record exists
                if m2m_fields_data:
                    from sqlalchemy import select as _sa_select

                    for rel_name, selected_ids in m2m_fields_data.items():
                        if not hasattr(created_record, rel_name):
                            continue
                        target_config = next(
                            (
                                c
                                for c in registry.all()
                                if hasattr(c.model, "__table__")
                                and c.model.__name__
                                == next(
                                    (
                                        f.target_model
                                        for f in fields
                                        if f.name == rel_name
                                    ),
                                    None,
                                )
                            ),
                            None,
                        )
                        if target_config and selected_ids:
                            res = await session.execute(
                                _sa_select(target_config.model).where(
                                    target_config.model.id.in_(
                                        [int(i) for i in selected_ids if i]
                                    )
                                )
                            )
                            related_objs = res.scalars().all()
                            setattr(created_record, rel_name, list(related_objs))
                        else:
                            setattr(created_record, rel_name, [])

                # Audit Log
                if audit_logger:
                    await audit_logger.log_create(
                        session,
                        model_name=model,
                        record_id=str(getattr(created_record, "id", "new")),
                        record_data=model_to_dict(
                            created_record,
                            exclude={"password_hash", "totp_secret"},
                        ),
                        user_id=getattr(current_user, "id", None),
                        username=getattr(current_user, "username", None),
                        ip_address=request.client.host if request.client else None,
                    )

                await session.commit()
            except Exception as e:
                await session.rollback()
                # Re-render form with error message
                fragment_token = signer.create_fragment_token(
                    model, action="load_fragment"
                )
                prefix = request.scope.get("root_path", "") + router.prefix
                fragment_url = f"{prefix}/fragments?token={fragment_token}"

                context = {
                    **get_common_context(request),
                    "model_config": model_config,
                    "fields": fields,
                    "values": create_data,
                    "errors": {"__all__": [str(e)]},
                    "record_id": None,
                    "fragment_url": fragment_url,
                    "csrf_token": signer.sign({"action": "create", "model": model}),
                    "current_user": current_user,
                    "detail_panels": model_config.detail_panels,
                    "delete_token": signer.sign({"model": model, "action": "delete"}),
                }
                return templates.TemplateResponse(request, "pages/edit.html", context)

        return RedirectResponse(
            url=str(request.url_for("admin:list", model=model)),
            status_code=303,
        )

    # ==================== Edit View ====================

    @router.get("/{model}/{id}", response_class=HTMLResponse, name="admin:edit")
    async def edit_view(
        request: Request,
        model: str,
        id: str,
        session: "AsyncSession" = (
            Depends(session_dependency) if session_dependency else None
        ),
    ):
        """Show edit form for a record."""
        try:
            model_config = registry.validate_model_access(model)
        except Exception:
            raise HTTPException(
                status_code=HTTP_403_FORBIDDEN, detail="Model not registered"
            )

        # Get form fields - use SQLAlchemy inspector for SQLAlchemy models
        if hasattr(model_config.model, "__tablename__"):
            # SQLAlchemy model
            fields = extract_sqlalchemy_fields(
                model_config.model,
                exclude=model_config.exclude,
                include=model_config.fields,
                registry=registry,
                field_overrides=model_config.field_overrides,
                widgets=model_config.widgets,
            )
        else:
            # Pydantic model
            fields = walker.walk(
                model_config.model,
                exclude=model_config.exclude,
                include=model_config.fields if model_config.fields else None,
            )

        # Fetch actual record data from database
        if session and hasattr(model_config.model, "__tablename__"):
            crud = CRUDBase(model_config.model)
            current_user = await enforce_permission(
                request, session, model_config, "view"
            )
            query_transform = scoped_query_transform(
                model_config, request, session, current_user
            )
            record = await crud.get(
                session,
                id,
                query_transform=query_transform,
                load_relationships=model_config.eager_load,
            )

            if not record:
                raise HTTPException(
                    status_code=HTTP_404_NOT_FOUND, detail="Record not found"
                )

            # Convert scalar columns to dict; collect M2M current IDs separately
            m2m_values: dict[str, list[str]] = {}
            values = {}
            for f in fields:
                if f.field_type == FieldType.MANY_TO_MANY:
                    related = getattr(record, f.name, []) or []
                    m2m_values[f.name] = [str(getattr(r, "id", "")) for r in related]
                else:
                    values[f.name] = getattr(record, f.name, None)
            values["id"] = id

            m2m_choices = await _load_m2m_choices(session, fields, registry)
        else:
            values = {"id": id}
            m2m_values = {}
            m2m_choices = {}

        # Generate signed fragment URL
        fragment_token = signer.create_fragment_token(
            model, action="load_fragment", record_id=id
        )
        fragment_url = f"{prefix}/fragments?token={fragment_token}"

        context = {
            **get_common_context(request),
            "model_config": model_config,
            "current_model": model,
            "fields": fields,
            "values": values,
            "errors": {},
            "record_id": id,
            "fragment_url": fragment_url,
            "csrf_token": signer.sign({"action": "update", "model": model, "id": id}),
            "current_user": locals().get("current_user"),
            "detail_panels": model_config.detail_panels,
            "record": locals().get("record"),
            "delete_token": signer.sign({"model": model, "action": "delete"}),
            "m2m_choices": m2m_choices,
            "m2m_values": m2m_values,
        }

        return templates.TemplateResponse(request, "pages/edit.html", context)

    @router.post("/{model}/{id}", name="admin:update")
    async def update_submit(
        request: Request,
        model: str,
        id: str,
        session: "AsyncSession" = (
            Depends(session_dependency) if session_dependency else None
        ),
    ):
        """Handle update form submission."""
        try:
            model_config = registry.validate_model_access(model)
        except Exception:
            raise HTTPException(
                status_code=HTTP_403_FORBIDDEN, detail="Model not registered"
            )

        if model_config.readonly:
            raise HTTPException(
                status_code=HTTP_403_FORBIDDEN, detail="Model is read-only"
            )

        current_user = await enforce_permission(request, session, model_config, "edit")
        query_transform = scoped_query_transform(
            model_config, request, session, current_user
        )

        form_data = await request.form()

        # Validate CSRF token
        csrf_token = form_data.get("_csrf_token", "")
        try:
            signer.unsign(csrf_token)
        except Exception:
            raise HTTPException(
                status_code=HTTP_403_FORBIDDEN, detail="Invalid CSRF token"
            )

        # Update record in database
        if session and hasattr(model_config.model, "__tablename__"):
            crud = CRUDBase(model_config.model)

            # Convert form data to dict (excluding internal fields)
            raw_data = {
                k: v for k, v in dict(form_data).items() if not k.startswith("_")
            }

            # Convert string values to appropriate types based on model
            update_data = {}
            m2m_fields_data: dict[str, list[str]] = {}
            fields = extract_sqlalchemy_fields(
                model_config.model,
                registry=registry,
                field_overrides=model_config.field_overrides,
                widgets=model_config.widgets,
            )
            for field in fields:
                if field.field_type == FieldType.MANY_TO_MANY:
                    m2m_fields_data[field.name] = form_data.getlist(field.name)
                    continue
                if field.name in raw_data:
                    value = raw_data[field.name]
                    # Convert boolean strings to actual booleans
                    if field.field_type == FieldType.BOOLEAN:
                        update_data[field.name] = value in (
                            "true",
                            "True",
                            "1",
                            "on",
                            True,
                        )
                    # Convert numeric strings to numbers
                    elif field.field_type == FieldType.NUMBER and isinstance(
                        value, str
                    ):
                        update_data[field.name] = int(value) if value else None
                    elif field.field_type == FieldType.FLOAT and isinstance(value, str):
                        update_data[field.name] = float(value) if value else None
                    elif field.field_type == FieldType.JSON and isinstance(value, str):
                        import json

                        update_data[field.name] = json.loads(value) if value else None
                    else:
                        update_data[field.name] = value if value != "" else None

            try:
                existing_record = await crud.get(
                    session,
                    id,
                    query_transform=query_transform,
                    load_relationships=model_config.eager_load,
                )
                old_data = (
                    model_to_dict(
                        existing_record, exclude={"password_hash", "totp_secret"}
                    )
                    if existing_record
                    else {}
                )

                updated_record = await crud.update(
                    session,
                    id=id,
                    obj_in=update_data,
                    query_transform=query_transform,
                )
                if not updated_record:
                    raise HTTPException(
                        status_code=HTTP_404_NOT_FOUND, detail="Record not found"
                    )

                # Update M2M relationships
                if m2m_fields_data:
                    from sqlalchemy import select as _sa_select

                    for rel_name, selected_ids in m2m_fields_data.items():
                        if not hasattr(updated_record, rel_name):
                            continue
                        target_config = next(
                            (
                                c
                                for c in registry.all()
                                if c.model.__name__
                                == next(
                                    (
                                        f.target_model
                                        for f in fields
                                        if f.name == rel_name
                                    ),
                                    None,
                                )
                            ),
                            None,
                        )
                        if target_config and selected_ids:
                            res = await session.execute(
                                _sa_select(target_config.model).where(
                                    target_config.model.id.in_(
                                        [int(i) for i in selected_ids if i]
                                    )
                                )
                            )
                            related_objs = res.scalars().all()
                            setattr(updated_record, rel_name, list(related_objs))
                        else:
                            setattr(updated_record, rel_name, [])

                # Audit Log
                if audit_logger:
                    await audit_logger.log_update(
                        session,
                        model_name=model,
                        record_id=id,
                        old_data=old_data,
                        new_data=model_to_dict(
                            updated_record,
                            exclude={"password_hash", "totp_secret"},
                        ),
                        user_id=getattr(current_user, "id", None),
                        username=getattr(current_user, "username", None),
                        ip_address=request.client.host if request.client else None,
                    )

                await session.commit()
            except HTTPException:
                raise
            except Exception as e:
                await session.rollback()
                raise HTTPException(status_code=400, detail=str(e))

        return RedirectResponse(
            url=str(request.url_for("admin:list", model=model)),
            status_code=303,
        )

    # ==================== Delete ====================

    @router.delete("/{model}/{id}", name="admin:delete")
    async def delete_record(
        request: Request,
        model: str,
        id: str,
        token: str = Query(...),
        session: "AsyncSession" = (
            Depends(session_dependency) if session_dependency else None
        ),
    ):
        """Delete a record."""
        # Validate signed token
        try:
            payload = signer.unsign(token)
            if payload.get("model") != model or payload.get("action") != "delete":
                raise ValueError("Token mismatch")
        except Exception:
            raise HTTPException(
                status_code=HTTP_403_FORBIDDEN, detail="Invalid or expired token"
            )

        try:
            model_config = registry.validate_model_access(model)
        except Exception:
            raise HTTPException(
                status_code=HTTP_403_FORBIDDEN, detail="Model not registered"
            )

        if model_config.readonly:
            raise HTTPException(
                status_code=HTTP_403_FORBIDDEN, detail="Model is read-only"
            )

        current_user = await enforce_permission(
            request, session, model_config, "delete"
        )
        query_transform = scoped_query_transform(
            model_config, request, session, current_user
        )

        # Delete record from database
        if session and hasattr(model_config.model, "__tablename__"):
            crud = CRUDBase(model_config.model)

            try:
                existing_record = await crud.get(
                    session,
                    id,
                    query_transform=query_transform,
                    load_relationships=model_config.eager_load,
                )
                deleted = await crud.delete(
                    session,
                    id=id,
                    query_transform=query_transform,
                )
                if not deleted:
                    raise HTTPException(
                        status_code=HTTP_404_NOT_FOUND, detail="Record not found"
                    )

                # Audit Log
                if audit_logger:
                    await audit_logger.log_delete(
                        session,
                        model_name=model,
                        record_id=id,
                        record_data=model_to_dict(
                            existing_record,
                            exclude={"password_hash", "totp_secret"},
                        ),
                        user_id=getattr(current_user, "id", None),
                        username=getattr(current_user, "username", None),
                        ip_address=request.client.host if request.client else None,
                    )

                await session.commit()
            except HTTPException:
                raise
            except Exception as e:
                await session.rollback()
                raise HTTPException(status_code=400, detail=str(e))

        # Return empty response for HTMX to remove the row
        return HTMLResponse(content="", status_code=200)

    # ==================== Polymorphic Fragment Loading ====================

    @router.get("/api/search/{model}", name="admin:search_related")
    async def search_related(
        request: Request,
        model: str,
        q: str = Query(""),
        session: "AsyncSession" = (
            Depends(session_dependency) if session_dependency else None
        ),
    ):
        """Search related models for relationship inputs."""
        if not session:
            return JSONResponse([])

        try:
            model_config = registry.validate_model_access(model)
        except Exception:
            raise HTTPException(
                status_code=HTTP_403_FORBIDDEN,
                detail="Model not registered",
            )

        current_user = await enforce_permission(request, session, model_config, "view")
        query_transform = scoped_query_transform(
            model_config, request, session, current_user
        )
        crud = CRUDBase(model_config.model)

        rows, _ = await crud.list(
            session,
            page=1,
            per_page=10,
            search=q,
            search_fields=model_config.searchable_fields or ["id"],
            order_by=model_config.ordering,
            query_transform=query_transform,
        )

        items = []
        for row in rows:
            label = (
                row.__admin_repr__()
                if hasattr(row, "__admin_repr__")
                else getattr(row, "name", None)
                or getattr(row, "title", None)
                or getattr(row, "email", None)
                or f"{model_config.name} #{getattr(row, 'id', '')}"
            )
            items.append({"id": getattr(row, "id", ""), "label": str(label)})
        return JSONResponse(items)

    @router.post("/{model}/actions/{action_name}", name="admin:run_action")
    async def run_action(
        request: Request,
        model: str,
        action_name: str,
        session: "AsyncSession" = (
            Depends(session_dependency) if session_dependency else None
        ),
    ):
        """Run a bulk or custom action for a model."""
        if not session:
            raise HTTPException(
                status_code=400, detail="Bulk actions require a database session"
            )

        try:
            model_config = registry.validate_model_access(model)
        except Exception:
            raise HTTPException(
                status_code=HTTP_403_FORBIDDEN,
                detail="Model not registered",
            )

        current_user = await enforce_permission(request, session, model_config, "edit")
        query_transform = scoped_query_transform(
            model_config, request, session, current_user
        )
        form_data = await request.form()
        selected_ids = form_data.getlist("selected_ids")
        crud = CRUDBase(model_config.model)

        if action_name == "bulk_delete":
            deleted = await crud.bulk_delete(
                session,
                ids=selected_ids,
                query_transform=query_transform,
            )
            if audit_logger and deleted:
                await audit_logger.log_delete(
                    session,
                    model_name=model,
                    record_id="bulk",
                    record_data={"deleted_ids": selected_ids, "count": deleted},
                    user_id=getattr(current_user, "id", None),
                    username=getattr(current_user, "username", None),
                    ip_address=request.client.host if request.client else None,
                )
            await session.commit()
            return RedirectResponse(
                url=str(request.url_for("admin:list", model=model)),
                status_code=303,
            )

        action = next(
            (item for item in model_config.actions if item.name == action_name), None
        )
        if action is None:
            raise HTTPException(
                status_code=HTTP_404_NOT_FOUND, detail="Action not found"
            )

        result = await action.handler(
            request=request,
            session=session,
            model=model_config.model,
            ids=selected_ids,
            user=current_user,
            config=model_config,
        )
        await session.commit()
        return render_action_result(
            result, str(request.url_for("admin:list", model=model))
        )

    @router.get("/fragments", response_class=HTMLResponse, name="admin:load_fragment")
    async def load_fragment(
        request: Request,
        token: str = Query(...),
        subtype: str = Query(None),
        session: "AsyncSession" = (
            Depends(session_dependency) if session_dependency else None
        ),
    ):
        """
        Load form fields for a polymorphic subtype.

        This endpoint is called via HTMX when the user selects a different
        type in a discriminated union dropdown. It uses signed tokens to
        prevent IDOR attacks.
        """
        # Validate the signed token
        try:
            payload = signer.unsign(token)
        except Exception as e:
            raise HTTPException(
                status_code=HTTP_403_FORBIDDEN,
                detail=f"Invalid or expired token: {e}",
            )

        # Extract and validate model
        model_name = payload.get("model")
        if not model_name:
            raise HTTPException(
                status_code=HTTP_403_FORBIDDEN, detail="Missing model in token"
            )

        try:
            model_config = registry.validate_model_access(model_name)
        except Exception:
            raise HTTPException(
                status_code=HTTP_403_FORBIDDEN, detail="Model not registered"
            )
        await enforce_permission(request, session, model_config, "view")

        # Get subtype from query or form
        if not subtype:
            form_data = await request.form()
            # Get from discriminator field value in the request
            for key, value in form_data.items():
                if value and key not in ["_model", "_csrf_token"]:
                    subtype = value
                    break

        if not subtype:
            return HTMLResponse(
                content="<p class='text-muted-foreground text-sm'>Select a type above</p>"
            )

        # Validate subtype access (Anti-Type-Confusion)
        try:
            subtype_class = registry.validate_subtype_access(model_name, subtype)
        except Exception:
            raise HTTPException(
                status_code=HTTP_403_FORBIDDEN,
                detail=f"Subtype '{subtype}' is not allowed for model '{model_name}'",
            )

        # Walk the subtype to get its fields
        # Find the discriminator field name
        discriminator = None
        parent_fields = walker.walk(model_config.model)
        for field in parent_fields:
            if field.discriminator:
                discriminator = field.discriminator
                break

        fields = walker.walk_subtype(subtype_class, parent_discriminator=discriminator)

        context = {
            "request": request,
            "fields": fields,
            "values": {},
            "errors": {},
            "model_name": model_name,
            "csp_nonce": getattr(request.state, "csp_nonce", ""),
        }

        return templates.TemplateResponse(
            request, "fragments/form_fields.html", context
        )

    return router
